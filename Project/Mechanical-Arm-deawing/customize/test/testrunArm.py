import serial
import lewansoul_lx16a
import time
import openpyxl
import numpy as np
import cv2
# import matplotlib.pyplot as plt
from project.sobel import sobel_edge_detection
from project.gaussian_smoothing import gaussian_blur
from project.myCannyEdge import non_max_suppression, hysteresis, threshold

#==============Setup =====================
dl =  1.3
dl2 = 2

wd = 35
ht = 100
prv_step = 0

ser = serial.Serial('COM9', 9600, timeout=1)

SERIAL_PORT = "COM12"

controller = lewansoul_lx16a.ServoController(serial.Serial(SERIAL_PORT, 115200, timeout=1),)
time.sleep(5)

#=========Setup Data============================
dataframe = openpyxl.load_workbook("test001.xlsx")

# print(dataframe.sheetnames[0])
dataframe1 = dataframe.worksheets[0]
m1 = np.zeros((ht, wd))
m2 = np.zeros((ht, wd))
m3 = np.zeros((ht, wd))
m4 = np.zeros((ht, wd))
m5 = np.zeros((ht, wd))

# stepMsT = np.array([1, 600, 1200, 1800, 2400, 3000, 3600, 4200, 4800, 5400, 6000, 6600, 7200, 7800, 8400, 9000, 9600, 10200, 10800, 11400, 12000, 12600, 13200, 13800, 14400, 15000, 15600, 16200, 16800, 17400, 18000, 18600, 19200, 19800, 20400, 21000,
# stepMsT = np.array([1, 300, 600, 900, 1200, 1500, 1800, 2100, 2400, 2700, 3000, 3300, 3600, 3900, 4200, 4500, 4800, 5100, 5400, 5700, 6000, 6300, 6600, 6900, 7200, 7500, 7800, 8100, 8400, 8700, 9000, 9300, 9600, 9900, 10200, 10500, 10800, 11100, 11400, 11700, 12000, 12300, 12600, 12900, 13200, 13500, 13800, 14100, 14400, 14700, 15000, 15300, 15600, 15900, 16200, 16500, 16800, 17100, 17400, 17700, 18000, 18300, 18600, 18900, 19200, 19500, 19800, 20100, 20400, 20700, 21000,
#                    # 21600,	22200,	22800,	23400,	24000,	24600,	25200,	25800,	26400,	27000,	27600,	28200,	28800,	29400,	30000,	30600,	31200,	31800,	32400,	33000,	33600,	34200,	34800,	35400,	36000,	36600,	37200,	37800,	38400,	39000,	39600,	40200,	40800,	41400,	42000
#                     ])
stepMsT = np.array([1, 450, 900, 1350,	1800, 2250,	2700, 3150,	3600, 4050,	4500, 4950,	5400, 5850,	6300, 6750,	7200, 7650,	8100, 8550,	9000, 9450,	9900, 10350, 10800,	11250, 11700, 12150, 12600,	13050, 13500, 13950, 14400, 14850, 15300, 15750
                   # 21600,	22200,	22800,	23400,	24000,	24600,	25200,	25800,	26400,	27000,	27600,	28200,	28800,	29400,	30000,	30600,	31200,	31800,	32400,	33000,	33600,	34200,	34800,	35400,	36000,	36600,	37200,	37800,	38400,	39000,	39600,	40200,	40800,	41400,	42000
                    ])
print(stepMsT.shape)

stepMs = stepMsT[1:wd+1]

mstep = np.zeros((ht, wd))

cntc = 0
for cole in dataframe1.iter_cols(1, 1):
    cntr = 0
    for row in range(0, ht):

        # for col in dataframe1.iter_cols(1, 2):

        mystr = cole[row].value

        spstr = cole[row].value

        x = spstr.split(",")

        print("x =" + x[0] + ", y =" + x[1] + ", z =" + x[2] + ", v =" + x[3] + ", sv =" + x[4] + ", step =" + x[5])
        for col in range(0, wd):
            m1[int(cntr), int(col)] = int(x[0])
            m2[int(cntr), int(col)] = int(x[1])
            m3[int(cntr), int(col)] = int(x[2])
            m4[int(cntr), int(col)] = int(x[3])
            m5[int(cntr), int(col)] = int(x[4])
            mstep[int(cntr), int(col)] = stepMs[col]
        cntr = cntr + 1

    cntc = cntc + 1

#==========Get Image======================

# image = cv2.imread("Image/black1.jpg")
image = cv2.imread("Image/test.png")
# image = cv2.imread("Image/test2.png")
image = np.array(image)
# print(image.shape)
# resize = cv2.resize(image, (115, 110))

resize = image
cv2.imshow("Resized Image", resize)

# plt.imshow(image, cmap='gray')
cv2.namedWindow("Original Image", cv2.WINDOW_NORMAL)
cv2.imshow("Original Image", image)
# plt.show()

blurred_image = gaussian_blur(image, kernel_size=9, verbose=False)

edge_filter = np.array([[-1, 0, 1], [-2, 0, 2], [-1, 0, 1]])

# gradient_magnitude, gradient_direction = sobel_edge_detection(blurred_image, edge_filter, convert_to_degree=True,
#                                                              verbose=args["verbose"])
gradient_magnitude, gradient_direction = sobel_edge_detection(blurred_image, edge_filter, convert_to_degree=True)

# new_image = non_max_suppression(gradient_magnitude, gradient_direction, verbose=args["verbose"])
new_image = non_max_suppression(gradient_magnitude, gradient_direction)

weak = 50

# new_image = threshold(new_image, 5, 20, weak=weak, verbose=args["verbose"])
new_image = threshold(new_image, 5, 20, weak=weak)
new_image = hysteresis(new_image, weak)
# new_image = image
cv2.namedWindow("Edge Image", cv2.WINDOW_NORMAL)
cv2.imshow("Edge Image", new_image)

newSize_image = cv2.resize(new_image, (2*wd, ht), cv2.INTER_NEAREST)
# newSize_image = new_image
cv2.namedWindow("New Edge Image", cv2.WINDOW_NORMAL)
cv2.imshow("New Edge Image", newSize_image)

cv2.waitKey(5)
posPix = np.array(np.nonzero(newSize_image[:, 0:wd]))

print(posPix.shape)
y0 = posPix[0, :]
x0 = posPix[1, :]

posPix1 = np.array(np.nonzero(newSize_image[:, wd:]))

print(posPix1.shape)
y1 = posPix1[0, :]
x1 = posPix1[1, :]

#=============Setup Motor=====================




prv_step = -1
stepM = 0

res = bytes('L' + str(stepM - prv_step), 'utf-8')
ser.write(res + b'\n')
# print(res)
prv_step = stepM
# print(prv_step)

# print(dataframe1)
time.sleep(3)

print(len(x0))
#==============First Half=========================
for i in range(len(x0)):
    print(x0[i], ', ', y0[i], ' = ', newSize_image[y0[i], x0[i]])
    Step = int(mstep[y0[i], x0[i]])

    if Step >= prv_step:
        myStr = str(abs(Step - prv_step))
        prv_step = Step
        pulse = Step
        res = bytes('L' + myStr, 'utf-8')
        ser.write(res + b'\n')
        print(res)
        print(prv_step)
        # print('s1')
    else:
        myStr = str(abs(Step - prv_step))
        prv_step = Step
        pulse = Step
        res = bytes('R' + myStr, 'utf-8')
        ser.write(res + b'\n')
        print(res)
        print(prv_step)
        # print('s2')

    # print('s22')
    prv_step = Step
    print(prv_step)
    time.sleep(1)

    controller.move(1, int(m1[y0[i], x0[i]]), 1000)
    time.sleep(dl)
    controller.move(4, int(m4[y0[i], x0[i]]), 1000)
    time.sleep(dl)
    controller.move(3, int(m3[y0[i], x0[i]]), 1000)
    time.sleep(dl)
    # print('s3')

    SV = int(m5[y0[i], x0[i]])
    # print(SV)
    myStr = str(SV)
    res = bytes('S' + myStr, 'utf-8')
    ser.write(res + b'\n')
    print(res)
    time.sleep(1)
    controller.move(2, int(m2[y0[i], x0[i]]), 1000)
    time.sleep(dl)

    # print('s4')

    controller.move(2, 10, 1000)
    time.sleep(dl2)

    SV = 1
    # print(SV)
    myStr = str(SV)
    res = bytes('S' + myStr, 'utf-8')
    ser.write(res + b'\n')
    print(res)
    time.sleep(1)
    # print('s5')

    controller.move(3, 300, 1000)
    time.sleep(dl2)

Step = 16200

myStr = str(abs(Step - prv_step))
prv_step = 0
pulse = 0
res = bytes('L' + myStr, 'utf-8')
ser.write(res + b'\n')
print(res)
print(prv_step)

# res = bytes('L' + str(prv_step), 'utf-8')
# ser.write(res + b'\n')
# prv_step = step
# print(res)

#==============Second Half==========================
for i in range(posPix1.shape[1]):
    print(x1[i], ', ', y1[i], ' = ', newSize_image[y1[i], x1[i]])
    Step = int(mstep[y1[i], x1[i]])

    if Step >= prv_step:
        myStr = str(abs(Step - prv_step))
        prv_step = Step
        pulse = Step
        res = bytes('L' + myStr, 'utf-8')
        ser.write(res + b'\n')
        print(res)
        print(prv_step)
        # print('s1')
    else:
        myStr = str(abs(Step - prv_step))
        prv_step = Step
        pulse = Step
        res = bytes('R' + myStr, 'utf-8')
        ser.write(res + b'\n')
        print(res)
        print(prv_step)
        # print('s2')

    # print('s22')
    prv_step = Step
    print(prv_step)
    time.sleep(1)

    controller.move(1, int(m1[y1[i], x1[i]]), 1000)
    time.sleep(dl)
    controller.move(4, int(m4[y1[i], x1[i]]), 1000)
    time.sleep(dl)
    controller.move(3, int(m3[y1[i], x1[i]]), 1000)
    time.sleep(dl)
    # print('s3')

    SV = int(m5[y1[i], x1[i]])
    # print(SV)
    myStr = str(SV)
    res = bytes('S' + myStr, 'utf-8')
    ser.write(res + b'\n')
    print(res)
    time.sleep(1)
    controller.move(2, int(m2[y1[i], x1[i]]), 1000)
    time.sleep(dl)

    # print('s4')

    controller.move(2, 10, 1000)
    time.sleep(dl2)

    SV = 1
    # print(SV)
    myStr = str(SV)
    res = bytes('S' + myStr, 'utf-8')
    ser.write(res + b'\n')
    print(res)
    time.sleep(1)
    # print('s5')

    controller.move(3, 300, 1000)
    time.sleep(dl2)


#==================Move To Begin===========================
Step = 1
myStr = str(abs(Step - prv_step))
prv_step = Step
pulse = Step
res = bytes('R' + myStr, 'utf-8')
ser.write(res + b'\n')
print(res)
print(prv_step)

# step = 1
# res = bytes('R' + str(prv_step), 'utf-8')
# ser.write(res + b'\n')
# prv_step = step
# print(res)




cv2.waitKey(10)
# while cv2.waitKey(1) & 0xFF != ord('q'):
#     pass

print("Finish")