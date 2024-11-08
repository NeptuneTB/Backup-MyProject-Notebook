import wx
import wx.xrc
import cv2
import numpy
import serial
import lewansoul_lx16a
import time
import openpyxl

ofs = 0
dl = 2
ser = serial.Serial('COM4', 9600, timeout=1)

SERIAL_PORT = "COM5"

controller = lewansoul_lx16a.ServoController(serial.Serial(SERIAL_PORT, 115200, timeout=1),)
time.sleep(5)

dataframe = openpyxl.load_workbook("current.xlsx")

print(dataframe.sheetnames[0])
dataframe1 = dataframe.worksheets[0]

controller.move(1, 350, 1000)
time.sleep(dl)
controller.move(2, 1, 1000)
time.sleep(dl)
controller.move(3, 500, 1000)
time.sleep(dl)
controller.move(4, 1, 1000)
time.sleep(dl)

controller.move(1, 350, 1000)
time.sleep(dl)
controller.move(2, 1, 1000)
time.sleep(dl)
controller.move(3, 500, 1000)
time.sleep(dl)
controller.move(4, 1, 1000)
time.sleep(dl)

SV = 5
myStr = str(SV)
res = bytes(myStr, 'utf-8')
ser.write(res + b'\n')
print(SV)
time.sleep(5)

for row in range(0, 9):
    mystr = ""
    print("row " + str(row + 1))
    for col in dataframe1.iter_cols(0, 9):
        mystr = col[row].value
        spstr = col[row].value
        x = spstr.split(",")
        print(mystr)
        print("SV1 =" + x[0] + ", SV2 =" + x[1] + ", SV3 =" + x[2] + ", SV4 =" + x[3] + ", SV5 =" + x[4] + ", Step0 =" + x[5])

        print(x[0])
        controller.move(1, int(x[0]), 1000)
        time.sleep(dl)
        print(x[1])
        controller.move(2, int(x[1]), 1000)
        time.sleep(dl)
        print(x[2])
        controller.move(3, int(x[2]), 1000)
        time.sleep(dl)
        print(x[3])
        controller.move(4, int(x[3]), 1000)
        time.sleep(dl)

        # Servo 5
        SV = int(x[4])-1
        print(SV)
        myStr = str(SV)
        res = bytes(myStr, 'utf-8')
        ser.write(res + b'\n')
        print(res)
        time.sleep(3)

        # Step0 (Servo 6)
        SV6_from_excel = int(x[5])
        myStr6 = str(SV6_from_excel)
        res6 = bytes(myStr6, 'utf-8')
        ser.write(res6 + b'\n')
        print(res6)
        time.sleep(3)

        # Reset Servo 5
        SV = 5
        print(SV)
        myStr = str(SV)
        res = bytes(myStr, 'utf-8')
        ser.write(res + b'\n')
        print(res)
        time.sleep(3)

        controller.move(2, 150, 1000)
        time.sleep(dl)
