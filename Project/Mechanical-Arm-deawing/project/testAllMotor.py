import serial
import lewansoul_lx16a
import time
import openpyxl

ofs = 0
dl =  1.3
dl2 = 2

# dl2 = 1.2

prv_step = 0

ser = serial.Serial('COM3', 9600, timeout=1)

SERIAL_PORT = "COM4"

controller = lewansoul_lx16a.ServoController(serial.Serial(SERIAL_PORT, 115200, timeout=1),)
time.sleep(5)

dataframe = openpyxl.load_workbook("test001.xlsx")

print(dataframe.sheetnames[0])
dataframe1 = dataframe.worksheets[0]

# stepM = 20001
#
# res = bytes('R' + str(stepM), 'utf-8')
# ser.write(res + b'\n')
# print(res)
# prv_step = stepM
# print(prv_step)

stepMs = [1, 600, 1200, 1800, 2400, 3000, 3600, 4200, 4800, 5400, 6000, 6600, 7200, 7800, 8400, 9000, 9600, 10200, 10800, 11400, 12000, 12600, 13200, 13800, 14400, 15000, 15600, 16200, 16800, 17400, 18000, 18600, 19200, 19800, 20400, 21000]
prv_step = 0
cnt = 0
for stepM in stepMs:
    # prv_step - stepM, 12000, 12600, 13200, 13800
    res = bytes('L' + str(stepM - prv_step), 'utf-8')
    ser.write(res + b'\n')
    print(res)
    prv_step = stepM
    print(prv_step)

    print(dataframe1)
    # time.sleep(1)

    for col in dataframe1.iter_cols(1, 2):
    # for row in range(0, 9):
        mystr = ""
        # print ("row " + str(row + 1))

        for row in range(0, 100):
        #for col in dataframe1.iter_cols(1, 2):
            mystr = col[row].value
            spstr = col[row].value
            x = spstr.split(",")
            print(mystr)
            print("x =" + x[0] + ", y =" + x[1] + ", z =" + x[2] + ", v =" + x[3] + ", sv =" + x[4] + ", step =" + x[5])
            print('s0')
            Step = int(stepMs[cnt])

            if Step >= prv_step:
                myStr = str(abs(Step - prv_step))
                prv_step = Step
                pulse = Step
                res = bytes('L' + myStr, 'utf-8')
                ser.write(res + b'\n')
                print(res)
                print(prv_step)
                print('s1')
            else:
                myStr = str(abs(Step - prv_step))
                prv_step = Step
                pulse = Step
                res = bytes('R' + myStr, 'utf-8')
                ser.write(res + b'\n')
                print(res)
                print(prv_step)
                print('s2')

            print('s22')

            time.sleep(1)

            controller.move(1, int(x[0]), 1000)
            time.sleep(dl)
            controller.move(4, int(x[3]), 1000)
            time.sleep(dl)
            controller.move(3, int(x[2]), 1000)
            time.sleep(dl)
            print('s3')

            SV = int(x[4])
            print(SV)
            myStr = str(SV)
            res = bytes('S' + myStr, 'utf-8')
            ser.write(res + b'\n')
            print(res)
            time.sleep(1)
            controller.move(2, int(x[1]), 1000)
            time.sleep(dl)

            print('s4')

            controller.move(2, 10, 1000)
            time.sleep(dl2)

            SV = 50
            print(SV)
            myStr = str(SV)
            res = bytes('S' + myStr, 'utf-8')
            ser.write(res + b'\n')
            print(res)
            time.sleep(1)
            print('s5')

            controller.move(3, 300, 1000)
            time.sleep(dl2)
            # controller.move(4, 500, 1000)
            # time.sleep(0.5)
            # controller.move(1, 423, 1000)
            # time.sleep(dl2)
            print('s7')

    cnt = cnt + 1
    prv_step = stepM
    print(prv_step)

step = 1
res = bytes('R' + str(prv_step), 'utf-8')
ser.write(res + b'\n')
prv_step = step
print(res)